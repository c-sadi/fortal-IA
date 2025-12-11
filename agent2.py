import os
import re
import base64
import datetime

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook, load_workbook

# --- CONFIGURATION ---
SCOPES = ['https://www.googleapis.com/auth/gmail.modify']
LABEL_NAME = "Traité"
EXCEL_FILE = "prospects.xlsx"

# --- Domaine interne à ignorer (ex : Google Calendar) ---
INTERNAL_DOMAIN = "calendar.google.com"


# ----------------------------------------------------------
# 🔐 Connexion Gmail + Création auto de token.json
# ----------------------------------------------------------
def gmail_service():
    creds = None

    # Charger token.json s'il existe
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)

    # Si pas de token → on le crée
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # ⚠️ Doit correspondre au credentials.json de ton nouveau projet
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES
            )
            creds = flow.run_local_server(port=0)

        # Sauvegarder token
        with open('token.json', 'w', encoding="utf-8") as token:
            token.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)


# --- Créer ou récupérer label ---
def get_or_create_label(service, label_name):
    try:
        labels = service.users().labels().list(userId='me').execute().get('labels', [])
        for label in labels:
            if label['name'] == label_name:
                return label['id']

        label = {
            'name': label_name,
            'labelListVisibility': 'labelShow',
            'messageListVisibility': 'show'
        }
        created_label = service.users().labels().create(userId='me', body=label).execute()
        return created_label['id']

    except HttpError as error:
        print(f"Erreur création label : {error}")
        return None


# --- Extraire infos depuis le corps ---
def extract_info(email_body):
    info = {}
    patterns = {
        'Prénom': r'Prénom\s*[:=]?\s*(.*)',
        'Nom': r'Nom\s*[:=]?\s*(.*)',
        'Email': r'Email\s*[:=]?\s*(.*)',
        'Téléphone': r'Téléphone\s*[:=]?\s*(.*)',
        'Adresse': r'Adresse\s*[:=]?\s*(.*)',
        'Ville': r'Ville\s*[:=]?\s*(.*)',
        'Code postal': r'Code postal\s*[:=]?\s*(.*)',
        'Département': r'Département\s*[:=]?\s*(.*)',
        'Bien recherché': r'Bien recherché\s*[:=]?\s*(.*)',
        'Budget': r"Budget d'achat\s*[:=]?\s*(.*)",
        'Financement': r'A un dossier de financement\s*[:=]?\s*(.*)',
        'Délai': r"Délai d'achat\s*[:=]?\s*(.*)",
        'Secteurs': r'Secteurs de recherche\s*[:=]?\s*(.*)',
        'Programme neuf': r'Est intéressé par du programme neuf\s*[:=]?\s*(.*)',
        'Jours disponibles': r'Jours disponibles\s*[:=]?\s*(.*)',
        'Plages horaires': r'Plages horaires\s*[:=]?\s*(.*)'
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, email_body, re.IGNORECASE)
        info[key] = match.group(1).strip() if match else ''

    info['Date_Entree'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return info


# --- Ajouter ou créer Excel ---
def save_to_excel(info_list):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(info_list[0].keys()))

    for info in info_list:
        ws.append(list(info.values()))

    wb.save(EXCEL_FILE)
    wb.close()
    print(f"Données sauvegardées dans {EXCEL_FILE}")


# --- Marquer email traité ---
def mark_email_processed(service, msg_id, label_id):
    try:
        service.users().messages().modify(
            userId='me',
            id=msg_id,
            body={'addLabelIds': [label_id], 'removeLabelIds': ['UNREAD']}
        ).execute()

        print(f"Email {msg_id} traité et marqué comme '{LABEL_NAME}'.")

    except HttpError as e:
        print(f"Erreur traitement email {msg_id}: {e}")


# ------------------------
# MAIN
# ------------------------
def main():
    service = gmail_service()
    label_id = get_or_create_label(service, LABEL_NAME)

    try:
        results = service.users().messages().list(
            userId='me',
            q="is:unread",
            maxResults=200
        ).execute()

        messages = results.get('messages', [])
        print(f"{len(messages)} emails non lus à traiter...")

        all_info = []

        for msg in messages:
            msg_id = msg['id']
            message = service.users().messages().get(
                userId='me', id=msg_id, format='full'
            ).execute()

            # ---- Filtrer confirmations Google Calendar ----
            headers = message['payload'].get('headers', [])
            subject = ""
            sender = ""

            for h in headers:
                if h['name'] == 'From':
                    sender = h['value']
                if h['name'] == 'Subject':
                    subject = h['value']

            if ("confirmation" in subject.lower()) or \
               ("invitation" in subject.lower()) or \
               (INTERNAL_DOMAIN in sender.lower()):
                continue

            # ---- Récupérer corps ----
            payload = message['payload']
            if 'parts' in payload:
                body_data = payload['parts'][0]['body'].get('data', '')
            else:
                body_data = payload['body'].get('data', '')

            if body_data:
                email_body = base64.urlsafe_b64decode(body_data.encode('ASCII')).decode('utf-8')
            else:
                email_body = ""

            info = extract_info(email_body)
            all_info.append(info)

            mark_email_processed(service, msg_id, label_id)

        if all_info:
            save_to_excel(all_info)

        print("Tous les emails clients ont été traités avec succès !")

    except HttpError as error:
        print(f"Une erreur est survenue : {error}")


if __name__ == "__main__":
    main()
