from __future__ import print_function
import os.path
import random
from datetime import datetime, timedelta, time as dtime, timezone
import base64
import openpyxl
from email.mime.text import MIMEText
import pandas as pd

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# ==============================
# CONFIGURATION
# ==============================
SCOPES_CALENDAR = ['https://www.googleapis.com/auth/calendar']
SCOPES_GMAIL = ['https://www.googleapis.com/auth/gmail.send']
SCOPES_DRIVE = ['https://www.googleapis.com/auth/drive']  # <-- AJOUT DRIVE

DUREE_APPEL_MIN = 30
TIMEZONE = 'Europe/Paris'

EXCEL_FILE = "prospects.xlsx"

# --- Liste des collaborateurs ---
collaborateurs = [
    {"nom": "Dupont", "prenom": "Marie", "email": "oussalahsalma9@gmail.com"},
    {"nom": "Martin", "prenom": "Alex", "email": "alex.pro@example.com"}
]

# ==============================
# SERVICES GOOGLE
# ==============================
def get_calendar_service():
    creds = None
    if os.path.exists('token_calendar.json'):
        creds = Credentials.from_authorized_user_file('token_calendar.json', SCOPES_CALENDAR)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES_CALENDAR)
            creds = flow.run_local_server(port=0)
        with open('token_calendar.json', 'w') as token:
            token.write(creds.to_json())

    return build('calendar', 'v3', credentials=creds)


def get_gmail_service():
    creds = None
    if os.path.exists('token_gmail.json'):
        creds = Credentials.from_authorized_user_file('token_gmail.json', SCOPES_GMAIL)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES_GMAIL)
            creds = flow.run_local_server(port=0)
        with open('token_gmail.json', 'w') as token:
            token.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)


# ----------- SERVICE DRIVE -----------
def get_drive_service():
    creds = None
    if os.path.exists('token_drive.json'):
        creds = Credentials.from_authorized_user_file('token_drive.json', SCOPES_DRIVE)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES_DRIVE)
            creds = flow.run_local_server(port=0)
        with open('token_drive.json', 'w') as token:
            token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)


# ==============================
# ENVOI D'EMAIL
# ==============================
def send_email(sender, to, subject, body_text):
    msg = MIMEText(body_text, "plain", "utf-8")
    msg['To'] = to
    msg['From'] = sender
    msg['Subject'] = subject

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    gmail_service.users().messages().send(
        userId="me",
        body={'raw': raw}
    ).execute()


# ==============================
# MAIL : Demande de documents
# ==============================
def send_documents_request_email(collab, prenom, nom, email):
    subject = "Documents nécessaires pour ton analyse patrimoniale"

    body = f"""
Bonjour {prenom},

Merci encore pour le temps accordé lors de notre échange téléphonique.
Comme convenu, je t’envoie la liste des documents nécessaires pour réaliser l’analyse gratuite de ta capacité d’achat.

Documents à transmettre :
- Pièce d’identité (recto/verso)
- Justificatif de domicile de moins de 3 mois
- 3 dernières fiches de paie
- 3 derniers relevés de compte bancaire
- 2 derniers avis d’imposition
- Contrat de travail
- Attestations d’épargne (livrets, assurance-vie, etc.)

Prochaine étape :
Une fois les documents reçus, je procède à ton analyse et je te recontacte pour un échange d'environ 30 minutes.

À très vite,
{collab['prenom']} {collab['nom']}
"""

    send_email(collab["email"], email, subject, body)


# ==============================
# DRIVE : Création du dossier client
# ==============================
def create_drive_folder(prenom, nom):
    folder_name = f"{prenom} {nom} - Documents"

    file_metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder"
    }

    folder = drive_service.files().create(
        body=file_metadata,
        fields="id"
    ).execute()

    print(f"Dossier Drive créé : {folder_name} (ID: {folder['id']})")
    return folder["id"]


# ==============================
# UTILITAIRES
# ==============================
def is_slot_free(service, start_dt, end_dt):
    start_dt = start_dt.replace(tzinfo=timezone(timedelta(hours=1)))
    end_dt = end_dt.replace(tzinfo=timezone(timedelta(hours=1)))

    events_result = service.events().list(
        calendarId='primary',
        timeMin=start_dt.isoformat(),
        timeMax=end_dt.isoformat(),
        singleEvents=True
    ).execute()

    return len(events_result.get('items', [])) == 0


def next_slot(start_dt):
    start_dt += timedelta(minutes=DUREE_APPEL_MIN)

    if start_dt.hour == 12:
        start_dt = start_dt.replace(hour=13, minute=0)

    if start_dt.hour >= 19:
        start_dt = start_dt + timedelta(days=1)
        start_dt = start_dt.replace(hour=10, minute=0)

    return start_dt


def parse_time_range(range_str):
    try:
        start_str, end_str = range_str.split('-')
        start_time = datetime.strptime(start_str.strip(), "%H:%M").time()
        end_time = datetime.strptime(end_str.strip(), "%H:%M").time()
        return start_time, end_time
    except:
        return dtime(10, 0), dtime(10, 30)


# ==============================
# MAIN
# ==============================
def main():
    global gmail_service, drive_service

    calendar_service = get_calendar_service()
    gmail_service = get_gmail_service()
    drive_service = get_drive_service()  # <-- AJOUT DRIVE

    print("Services Google Calendar + Gmail + Drive OK")

    if not os.path.exists(EXCEL_FILE):
        print(f"⚠️ Fichier {EXCEL_FILE} introuvable. Création d'un fichier vide.")
        # On crée un DataFrame avec les colonnes attendues (ajustez selon vos besoins)
        df_vide = pd.DataFrame(columns=[
            "Date réception", "Prénom", "Nom", "Email", "Téléphone", 
            "Sujet", "Résumé", "Jours disponibles", "Plages horaires", 
            "Date RDV", "Heure RDV", "Traité"
        ])
        # On le sauvegarde pour créer le fichier physique
        df_vide.to_excel(EXCEL_FILE, index=False)

    # Maintenant on peut charger le fichier sans erreur
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    if "Traité" not in headers:
        ws.cell(row=1, column=len(headers)+1, value="Traité")
        headers.append("Traité")

    traited_col_idx = headers.index("Traité") + 1

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):

        if row[traited_col_idx-1].value == "✔️":
            continue

        prenom = row[headers.index("Prénom")].value or "Client"
        nom = row[headers.index("Nom")].value or ""
        email = row[headers.index("Email")].value or ""
        plages_horaire = row[headers.index("Plages horaires")].value or "10:00-10:30"

        collab = random.choice(collaborateurs)

        # --- 1) Envoi de la demande de documents ---
        send_documents_request_email(collab, prenom, nom, email)

        # --- 2) Création automatique du dossier Drive ---
        create_drive_folder(prenom, nom)

        # --- 3) Gestion du RDV ---
        for ph in plages_horaire.split(','):
            start_t, end_t = parse_time_range(ph)

            start_dt = datetime.combine(
                datetime.now().date() + timedelta(days=1), start_t
            )
            end_dt = datetime.combine(
                datetime.now().date() + timedelta(days=1), end_t
            )

            while not is_slot_free(calendar_service, start_dt, end_dt):
                start_dt = next_slot(start_dt)
                end_dt = start_dt + timedelta(minutes=DUREE_APPEL_MIN)

            event = {
                'summary': f"RDV avec {prenom} {nom}",
                'description': f"Email: {email}",
                'start': {'dateTime': start_dt.isoformat(), 'timeZone': TIMEZONE},
                'end': {'dateTime': end_dt.isoformat(), 'timeZone': TIMEZONE},
                'attendees': [{'email': email}, {'email': collab['email']}]
            }

            calendar_service.events().insert(
                calendarId='primary',
                body=event,
                sendUpdates='all'
            ).execute()

            print(f"RDV créé pour {prenom} {nom}")

            ws.cell(row=i, column=traited_col_idx, value="✔️")
            break

    wb.save(EXCEL_FILE)
    wb.close()
    print("Tout est terminé.")


if __name__ == "__main__":
    main()
